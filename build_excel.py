"""
ClearMetric Rent vs Buy Calculator — Premium Excel Template
Product for Gumroad ($12.99)

4 Sheets:
  1. Rent vs Buy Calculator — inputs left (teal), results right (break-even, verdict, monthly comparison)
  2. Year-by-Year Comparison — 30 rows: costs, equity, investments, net wealth
  3. What-If Scenarios — 3 scenarios: current market, optimistic, conservative
  4. How To Use — instructions for each input and interpreting results

Design: Teal palette (Airbnb-style: #117A65 primary, #0E6655 dark, #D1F2EB input cells)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.formatting.rule import CellIsRule, FormulaRule
import os

# ============================================================
# DESIGN SYSTEM — Teal (Airbnb-style)
# ============================================================
TEAL = "117A65"
DARK_TEAL = "0E6655"
WHITE = "FFFFFF"
INPUT_TEAL = "D1F2EB"
LIGHT_GRAY = "F5F6FA"
MED_GRAY = "D5D8DC"
DARK_GRAY = "5D6D7E"
GREEN = "27AE60"
LIGHT_GREEN = "EAFAF1"
RED = "E74C3C"
LIGHT_RED = "FDEDEC"
YELLOW = "F39C12"
LIGHT_YELLOW = "FEF9E7"
ACCENT = "1ABC9C"
LIGHT_TEAL = "E8F6F3"

FONT_TITLE = Font(name="Calibri", size=20, bold=True, color=WHITE)
FONT_SUBTITLE = Font(name="Calibri", size=12, color="A3E4D7", italic=True)
FONT_SECTION = Font(name="Calibri", size=13, bold=True, color=WHITE)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LABEL = Font(name="Calibri", size=11, color="2C3E50")
FONT_INPUT = Font(name="Calibri", size=12, color=TEAL, bold=True)
FONT_VALUE = Font(name="Calibri", size=11, color="2C3E50")
FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=TEAL)
FONT_SMALL = Font(name="Calibri", size=9, color=DARK_GRAY, italic=True)
FONT_BIG = Font(name="Calibri", size=28, bold=True, color=WHITE)
FONT_BIG_LABEL = Font(name="Calibri", size=12, bold=True, color="A3E4D7")
FONT_GREEN = Font(name="Calibri", size=11, bold=True, color=GREEN)
FONT_RED = Font(name="Calibri", size=11, bold=True, color=RED)
FONT_WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_CTA = Font(name="Calibri", size=12, bold=True, color=TEAL)

FILL_TEAL = PatternFill(start_color=TEAL, end_color=TEAL, fill_type="solid")
FILL_DARK = PatternFill(start_color=DARK_TEAL, end_color=DARK_TEAL, fill_type="solid")
FILL_INPUT = PatternFill(start_color=INPUT_TEAL, end_color=INPUT_TEAL, fill_type="solid")
FILL_GRAY = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type="solid")
FILL_WHITE = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
FILL_LIGHT = PatternFill(start_color=LIGHT_TEAL, end_color=LIGHT_TEAL, fill_type="solid")
FILL_GREEN = PatternFill(start_color=LIGHT_GREEN, end_color=LIGHT_GREEN, fill_type="solid")
FILL_RED = PatternFill(start_color=LIGHT_RED, end_color=LIGHT_RED, fill_type="solid")
FILL_YELLOW = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")

THIN = Border(
    left=Side("thin", MED_GRAY), right=Side("thin", MED_GRAY),
    top=Side("thin", MED_GRAY), bottom=Side("thin", MED_GRAY),
)
ALIGN_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_L = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_R = Alignment(horizontal="right", vertical="center")


def header_bar(ws, row, c1, c2, text, fill=None):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    cell = ws.cell(row=row, column=c1, value=text)
    cell.font = FONT_SECTION
    cell.fill = fill or FILL_TEAL
    cell.alignment = ALIGN_C
    for c in range(c1, c2 + 1):
        ws.cell(row=row, column=c).fill = fill or FILL_TEAL
        ws.cell(row=row, column=c).border = THIN


def label_input(ws, row, lc, vc, label, value=None, fmt=None, hint=None):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=value)
    cv.font = FONT_INPUT
    cv.fill = FILL_INPUT
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt
    if hint:
        ch = ws.cell(row=row, column=vc + 1, value=hint)
        ch.font = FONT_SMALL
        ch.alignment = ALIGN_L


def label_calc(ws, row, lc, vc, label, formula, fmt=None, bold=False):
    cl = ws.cell(row=row, column=lc, value=label)
    cl.font = FONT_LABEL
    cl.fill = FILL_GRAY
    cl.border = THIN
    cl.alignment = ALIGN_L
    cv = ws.cell(row=row, column=vc, value=formula)
    cv.font = FONT_BOLD if bold else FONT_VALUE
    cv.fill = FILL_WHITE
    cv.border = THIN
    cv.alignment = ALIGN_R
    if fmt:
        cv.number_format = fmt


def cols(ws, widths):
    for letter, w in widths.items():
        ws.column_dimensions[letter].width = w


# ============================================================
# SHEET 1: RENT VS BUY CALCULATOR
# ============================================================
def build_rent_vs_buy(ws):
    ws.title = "Rent vs Buy Calculator"
    ws.sheet_properties.tabColor = TEAL
    cols(ws, {
        "A": 2, "B": 36, "C": 18, "D": 18, "E": 4,
        "F": 36, "G": 18, "H": 2,
    })

    for r in range(1, 75):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    title = ws.cell(row=2, column=2, value="RENT VS BUY CALCULATOR")
    title.font = FONT_TITLE
    title.alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.row_dimensions[3].height = 22
    sub = ws.cell(row=3, column=2, value="Compare renting vs buying. Enter your numbers in the teal cells.")
    sub.font = FONT_SUBTITLE
    sub.alignment = ALIGN_C

    # ===== LEFT: INPUTS =====
    header_bar(ws, 5, 2, 4, "PURCHASE DETAILS")
    label_input(ws, 6, 2, 3, "Home Purchase Price", 400000, "$#,##0")
    label_input(ws, 7, 2, 3, "Down Payment %", 0.20, "0%")
    label_input(ws, 8, 2, 3, "Mortgage Interest Rate", 0.065, "0.00%")
    label_input(ws, 9, 2, 3, "Mortgage Term (years)", 30, "0")

    header_bar(ws, 11, 2, 4, "RENTING")
    label_input(ws, 12, 2, 3, "Monthly Rent", 2000, "$#,##0")
    label_input(ws, 13, 2, 3, "Annual Rent Increase %", 0.03, "0.0%")

    header_bar(ws, 15, 2, 4, "OWNERSHIP COSTS")
    label_input(ws, 16, 2, 3, "Property Tax Rate %", 0.012, "0.00%")
    label_input(ws, 17, 2, 3, "Home Insurance ($/year)", 1500, "$#,##0")
    label_input(ws, 18, 2, 3, "HOA Fees ($/month)", 0, "$#,##0")
    label_input(ws, 19, 2, 3, "Maintenance % of value/year", 0.01, "0.0%")
    label_input(ws, 20, 2, 3, "Closing Costs %", 0.03, "0.0%")

    header_bar(ws, 22, 2, 4, "ASSUMPTIONS")
    label_input(ws, 23, 2, 3, "Home Appreciation Rate %", 0.035, "0.0%")
    label_input(ws, 24, 2, 3, "Investment Return %", 0.07, "0.0%", "opportunity cost")
    label_input(ws, 25, 2, 3, "Income Tax Bracket %", 0.24, "0.0%", "for mortgage deduction")
    label_input(ws, 26, 2, 3, "Years to Compare", 10, "0")

    # ===== RIGHT: RESULTS =====
    header_bar(ws, 5, 6, 7, "RESULTS", FILL_DARK)

    # Loan amount
    label_calc(ws, 6, 6, 7, "Loan Amount", "=C6*(1-C7)", "$#,##0")
    # Monthly P&I
    label_calc(ws, 7, 6, 7, "Monthly P&I",
               "=IF(C8>0,PMT(C8/12,C9*12,-G6),0)", "$#,##0", bold=True)
    # Down payment
    label_calc(ws, 8, 6, 7, "Down Payment", "=C6*C7", "$#,##0")
    # Closing costs
    label_calc(ws, 9, 6, 7, "Closing Costs", "=C6*C20", "$#,##0")

    # Monthly ownership cost (year 1)
    label_calc(ws, 11, 6, 7, "Monthly Tax+Ins+HOA+Maint",
               "=C6*C16/12+C17/12+C18+C6*C19/12", "$#,##0")
    label_calc(ws, 12, 6, 7, "Total Monthly (Buy)",
               "=G7+G11", "$#,##0", bold=True)

    # Verdict (from Year-by-Year summary)
    header_bar(ws, 14, 6, 7, "VERDICT")
    ws.merge_cells("F15:G18")
    for r in range(15, 19):
        for c in range(6, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
            ws.cell(row=r, column=c).border = THIN
    verdict_cell = ws.cell(row=15, column=6)
    verdict_cell.value = "='Year-by-Year'!K40"
    verdict_cell.font = Font(name="Calibri", size=14, bold=True, color="A3E4D7")
    verdict_cell.alignment = ALIGN_C

    # Monthly comparison
    header_bar(ws, 21, 6, 7, "MONTHLY COMPARISON")
    label_calc(ws, 22, 6, 7, "Buy (total)", "=G12", "$#,##0")
    label_calc(ws, 23, 6, 7, "Rent (year 1)", "=C12", "$#,##0")

    # Protection
    ws.protection.sheet = True
    ws.protection.set_password("")
    input_cells = [(6, 3), (7, 3), (8, 3), (9, 3), (12, 3), (13, 3),
                   (16, 3), (17, 3), (18, 3), (19, 3), (20, 3),
                   (23, 3), (24, 3), (25, 3), (26, 3)]
    for r, c in input_cells:
        ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 2: YEAR-BY-YEAR COMPARISON
# ============================================================
def build_year_by_year(wb):
    ws = wb.create_sheet("Year-by-Year")
    ws.sheet_properties.tabColor = "1ABC9C"
    fc = "'Rent vs Buy Calculator'"
    cols(ws, {
        "A": 2, "B": 6, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14, "I": 14, "J": 14, "K": 12, "L": 28,
    })

    for r in range(1, 45):
        for c in range(1, 13):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    # Title
    for r in range(1, 4):
        for c in range(2, 12):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:K1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:K2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="YEAR-BY-YEAR COMPARISON").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:K3")
    ws.cell(row=3, column=2, value="30-year projection. Buyer equity vs Renter investments.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Headers
    headers = ["Year", "Home Value", "Mortgage Bal", "Buyer Equity", "Cum Cost Buy", "Cum Cost Rent",
               "Renter Inv", "Buy Wins?", "Verdict"]
    for i, h in enumerate(headers):
        cell = ws.cell(row=5, column=2 + i, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_TEAL
        cell.alignment = ALIGN_C
        cell.border = THIN

    # Build 30 rows with formulas
    # We need: loan amount, monthly rate, term, appreciation, rent, rent growth, etc.
    # Excel formulas for amortization: remaining balance = FV(rate/12, year*12, PMT, -loan)
    # Home value = price * (1+appreciation)^year
    # This gets complex. We'll use a simplified structure with key formulas.

    for yr in range(30):
        r = 6 + yr
        ws.row_dimensions[r].height = 20

        # Year
        ws.cell(row=r, column=2, value=yr + 1).font = FONT_VALUE
        ws.cell(row=r, column=2).alignment = ALIGN_C
        ws.cell(row=r, column=2).border = THIN
        ws.cell(row=r, column=2).fill = FILL_GRAY

        # Home Value = Price * (1+appr)^year
        ws.cell(row=r, column=3, value=f"={fc}!C6*(1+{fc}!C23)^{yr+1}").font = FONT_VALUE
        ws.cell(row=r, column=3).number_format = "$#,##0"
        ws.cell(row=r, column=3).border = THIN

        # Mortgage Balance = FV of loan after (yr+1)*12 payments (remaining balance)
        # FV(rate, nper, -pmt, pv) = balance; pmt negative for payment out
        pmt_ref = f"{fc}!G7"
        loan_ref = f"{fc}!G6"
        rate_ref = f"{fc}!C8"
        ws.cell(row=r, column=4,
                value=f"=MAX(0,FV({rate_ref}/12,({yr}+1)*12,-{pmt_ref},{loan_ref}))").font = FONT_VALUE
        ws.cell(row=r, column=4).number_format = "$#,##0"
        ws.cell(row=r, column=4).border = THIN

        # Buyer Equity = Home Value - Mortgage Balance
        ws.cell(row=r, column=5, value=f"=C{r}-D{r}").font = FONT_BOLD
        ws.cell(row=r, column=5).number_format = "$#,##0"
        ws.cell(row=r, column=5).border = THIN

        # Cumulative Cost Buy: down + closing + sum(annual costs - tax savings)
        # Annual cost = P&I*12 + prop_tax + ins + HOA*12 + maint
        # Tax savings = CUMIPMT (interest paid) * tax_bracket
        if yr == 0:
            ws.cell(row=r, column=6,
                    value=f"={fc}!G8+{fc}!G9+{fc}!G7*12+{fc}!C6*{fc}!C16+{fc}!C17+{fc}!C18*12+{fc}!C6*{fc}!C19-CUMIPMT({fc}!C8/12,{fc}!C9*12,{fc}!G6,1,12,0)*{fc}!C25")
        else:
            int_start = yr * 12 + 1
            int_end = (yr + 1) * 12
            ws.cell(row=r, column=6,
                    value=f"=F{r-1}+{fc}!G7*12+{fc}!C6*(1+{fc}!C23)^{yr}*{fc}!C16+{fc}!C17+{fc}!C18*12+{fc}!C6*(1+{fc}!C23)^{yr}*{fc}!C19-CUMIPMT({fc}!C8/12,{fc}!C9*12,{fc}!G6,{int_start},{int_end},0)*{fc}!C25")
        ws.cell(row=r, column=6).number_format = "$#,##0"
        ws.cell(row=r, column=6).border = THIN

        # Cumulative Cost Rent
        if yr == 0:
            ws.cell(row=r, column=7, value=f"={fc}!C12*12").font = FONT_VALUE
        else:
            ws.cell(row=r, column=7, value=f"=G{r-1}+{fc}!C12*12*(1+{fc}!C13)^{yr}").font = FONT_VALUE
        ws.cell(row=r, column=7).number_format = "$#,##0"
        ws.cell(row=r, column=7).border = THIN

        # Renter Investments (simplified: down payment * (1+r)^year + savings invested)
        # Simplified: = Down * (1+inv)^year
        ws.cell(row=r, column=8, value=f"={fc}!G8*(1+{fc}!C24)^{yr+1}").font = FONT_VALUE
        ws.cell(row=r, column=8).number_format = "$#,##0"
        ws.cell(row=r, column=8).border = THIN

        # Buy Wins? (1 if equity > renter inv)
        ws.cell(row=r, column=9, value=f"=IF(E{r}>H{r},1,0)").font = FONT_VALUE
        ws.cell(row=r, column=9).border = THIN

        # Verdict text
        ws.cell(row=r, column=10, value=f'=IF(E{r}>H{r},"Buy wins","Rent wins")').font = FONT_BOLD
        ws.cell(row=r, column=10).border = THIN

        if yr % 2 == 1:
            for c in range(3, 11):
                ws.cell(row=r, column=c).fill = FILL_GRAY

        # Conditional: green when buy wins
        ws.conditional_formatting.add(
            f"E{r}:H{r}",
            FormulaRule(formula=[f"E{r}>H{r}"], fill=FILL_GREEN, font=FONT_GREEN))

    # Verdict in row 40 (summary)
    ws.cell(row=40, column=2, value="Summary").font = FONT_SECTION
    ws.cell(row=40, column=2).fill = FILL_TEAL
    ws.cell(row=40, column=2).alignment = ALIGN_C
    ws.merge_cells("B40:C40")
    ws.cell(row=40, column=11,
            value="=IF(COUNTIF(I6:I35,1)>0,\"Buy better after year \"&MATCH(1,I6:I35,0)&\"\",\"Renting better for \"&'Rent vs Buy Calculator'!C26&\" years\")").font = FONT_CTA
    ws.merge_cells("K40:L40")

    # Chart: Equity vs Renter Investments
    chart = LineChart()
    chart.title = "Buyer Equity vs Renter Investments"
    chart.style = 10
    chart.y_axis.title = "Value ($)"
    chart.y_axis.numFmt = "$#,##0"
    chart.x_axis.title = "Year"

    data_equity = Reference(ws, min_col=5, min_row=5, max_row=35)
    data_renter = Reference(ws, min_col=8, min_row=5, max_row=35)
    cats = Reference(ws, min_col=2, min_row=6, max_row=35)

    chart.add_data(data_equity, titles_from_data=True)
    chart.add_data(data_renter, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.line.solidFill = TEAL
    chart.series[0].graphicalProperties.line.width = 28000
    chart.series[1].graphicalProperties.line.solidFill = "27AE60"
    chart.series[1].graphicalProperties.line.width = 20000

    chart.width = 24
    chart.height = 14
    ws.add_chart(chart, "B42")


# ============================================================
# SHEET 3: WHAT-IF SCENARIOS
# ============================================================
def build_what_if(wb):
    ws = wb.create_sheet("What-If Scenarios")
    ws.sheet_properties.tabColor = "2E86C1"
    fc = "'Rent vs Buy Calculator'"
    cols(ws, {
        "A": 2, "B": 32, "C": 18, "D": 4, "E": 18, "F": 4, "G": 18, "H": 2,
    })

    for r in range(1, 50):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = FILL_WHITE

    for r in range(1, 4):
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = FILL_DARK
    ws.merge_cells("B1:G1")
    ws.row_dimensions[1].height = 10
    ws.merge_cells("B2:G2")
    ws.row_dimensions[2].height = 38
    ws.cell(row=2, column=2, value="WHAT-IF SCENARIOS").font = FONT_TITLE
    ws.cell(row=2, column=2).alignment = ALIGN_C
    ws.merge_cells("B3:G3")
    ws.cell(row=3, column=2, value="Compare current market vs optimistic vs conservative assumptions.").font = FONT_SUBTITLE
    ws.cell(row=3, column=2).alignment = ALIGN_C

    # Headers
    for c in range(2, 8):
        ws.cell(row=5, column=c).fill = FILL_TEAL
        ws.cell(row=5, column=c).border = THIN
    ws.cell(row=5, column=2, value="Parameter").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=2).alignment = ALIGN_C
    ws.cell(row=5, column=3, value="Current Market").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=3).alignment = ALIGN_C
    ws.cell(row=5, column=5, value="Optimistic").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=5).alignment = ALIGN_C
    ws.cell(row=5, column=7, value="Conservative").font = FONT_WHITE_BOLD
    ws.cell(row=5, column=7).alignment = ALIGN_C

    params = [
        (6, "Mortgage Rate", f"={fc}!C8", 0.055, 0.075, "0.00%"),
        (7, "Rent Increase", f"={fc}!C13", 0.02, 0.05, "0.0%"),
        (8, "Home Appreciation", f"={fc}!C23", 0.05, 0.02, "0.0%"),
        (9, "Investment Return", f"={fc}!C24", 0.08, 0.05, "0.0%"),
    ]

    for r, label, val_a, val_b, val_c, fmt in params:
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN

        ca = ws.cell(row=r, column=3, value=val_a)
        ca.font = FONT_BOLD
        ca.fill = FILL_LIGHT
        ca.border = THIN
        ca.alignment = ALIGN_C
        ca.number_format = fmt

        cb = ws.cell(row=r, column=5, value=val_b)
        cb.font = FONT_INPUT
        cb.fill = FILL_INPUT
        cb.border = THIN
        cb.alignment = ALIGN_C
        cb.number_format = fmt

        cc = ws.cell(row=r, column=7, value=val_c)
        cc.font = FONT_INPUT
        cc.fill = FILL_INPUT
        cc.border = THIN
        cc.alignment = ALIGN_C
        cc.number_format = fmt

    # Results
    header_bar(ws, 11, 2, 7, "RESULTS AT YEAR 10")

    def _res(r, label, formula_a, formula_b, formula_c, fmt):
        ws.row_dimensions[r].height = 22
        ws.cell(row=r, column=2, value=label).font = FONT_LABEL
        ws.cell(row=r, column=2).fill = FILL_GRAY
        ws.cell(row=r, column=2).border = THIN
        for col, formula in [(3, formula_a), (5, formula_b), (7, formula_c)]:
            cell = ws.cell(row=r, column=col, value=formula)
            cell.font = FONT_BOLD
            cell.fill = FILL_WHITE
            cell.border = THIN
            cell.alignment = ALIGN_C
            if fmt:
                cell.number_format = fmt

    # Buyer equity at year 10: Home value - remaining mortgage balance
    # FV(rate, 120, -PMT, pv) = balance after 120 payments (Excel: pmt negative for payment out)
    ws.cell(row=12, column=2, value="Buyer Equity (Yr 10)").font = FONT_LABEL
    ws.cell(row=12, column=2).fill = FILL_GRAY
    ws.cell(row=12, column=2).border = THIN
    ws.cell(row=12, column=3, value=f"={fc}!C6*(1+{fc}!C23)^10-FV({fc}!C8/12,120,-{fc}!G7,{fc}!G6)").font = FONT_BOLD
    ws.cell(row=12, column=3).number_format = "$#,##0"
    ws.cell(row=12, column=3).fill = FILL_WHITE
    ws.cell(row=12, column=3).border = THIN
    ws.cell(row=12, column=3).alignment = ALIGN_C
    ws.cell(row=12, column=5, value=f"={fc}!C6*(1+E8)^10-FV(E6/12,120,-{fc}!G7,{fc}!G6)").font = FONT_BOLD
    ws.cell(row=12, column=5).number_format = "$#,##0"
    ws.cell(row=12, column=5).fill = FILL_WHITE
    ws.cell(row=12, column=5).border = THIN
    ws.cell(row=12, column=5).alignment = ALIGN_C
    ws.cell(row=12, column=7, value=f"={fc}!C6*(1+G8)^10-FV(G6/12,120,-{fc}!G7,{fc}!G6)").font = FONT_BOLD
    ws.cell(row=12, column=7).number_format = "$#,##0"
    ws.cell(row=12, column=7).fill = FILL_WHITE
    ws.cell(row=12, column=7).border = THIN
    ws.cell(row=12, column=7).alignment = ALIGN_C

    _res(13, "Renter Investments (Yr 10)",
         f"={fc}!G8*(1+{fc}!C24)^10",
         f"={fc}!G8*(1+E9)^10",
         f"={fc}!G8*(1+G9)^10",
         "$#,##0")

    _res(14, "Winner",
         '=IF(C12>C13,"Buy","Rent")',
         '=IF(E12>E13,"Buy","Rent")',
         '=IF(G12>G13,"Buy","Rent")',
         None)

    ws.protection.sheet = True
    for r in range(6, 10):
        for c in [5, 7]:
            ws.cell(row=r, column=c).protection = openpyxl.styles.Protection(locked=False)


# ============================================================
# SHEET 4: HOW TO USE
# ============================================================
def build_instructions(wb):
    ws = wb.create_sheet("How To Use")
    ws.sheet_properties.tabColor = DARK_GRAY
    cols(ws, {"A": 3, "B": 90})

    ws.merge_cells("A1:B2")
    c = ws.cell(row=1, column=1, value="HOW TO USE THE RENT VS BUY CALCULATOR")
    c.font = FONT_TITLE
    c.fill = FILL_DARK
    c.alignment = ALIGN_C
    for r in range(1, 3):
        for co in range(1, 3):
            ws.cell(row=r, column=co).fill = FILL_DARK

    sections = [
        ("QUICK START", [
            "1. Open the 'Rent vs Buy Calculator' tab and enter your numbers in the TEAL cells",
            "2. Results appear on the right: monthly payment, break-even, verdict",
            "3. Check the 'Year-by-Year' tab for the full 30-year comparison",
            "4. Use 'What-If Scenarios' to test different market assumptions",
        ]),
        ("INPUT EXPLANATIONS", [
            "Home Purchase Price: The total price of the home you're considering",
            "Down Payment %: Typically 20% to avoid PMI; 3-5% for FHA/conventional low-down",
            "Mortgage Rate: Current market rate for your credit profile and term",
            "Mortgage Term: 30-year = lower payment, more interest; 15-year = higher payment, less interest",
            "Monthly Rent: What you'd pay to rent a comparable home",
            "Rent Increase %: Historical average ~3%/year; adjust for your market",
            "Property Tax %: Varies by location; 1-2% typical in US",
            "Home Insurance: Annual premium; get a quote for accuracy",
            "HOA Fees: Monthly HOA/condo fees; $0 for single-family",
            "Maintenance %: Rule of thumb 1% of home value per year",
            "Closing Costs %: Typically 2-5% of purchase price",
            "Appreciation %: Historical US average ~3.5%; varies by market",
            "Investment Return %: Long-term stock market ~7%; adjust for your allocation",
            "Tax Bracket %: For mortgage interest deduction benefit",
        ]),
        ("INTERPRETING RESULTS", [
            "Break-even year: When buyer equity exceeds renter investments",
            "Before break-even: Renting and investing the difference is better",
            "After break-even: Owning builds more wealth",
            "Verdict: Summary of which path wins for your timeframe",
            "Monthly comparison: Year 1 costs; rent grows over time, mortgage (P&I) is fixed",
        ]),
        ("WHAT-IF SCENARIOS", [
            "Current Market: Uses your main calculator inputs",
            "Optimistic: Lower rates, higher appreciation, higher investment returns",
            "Conservative: Higher rates, lower appreciation, lower investment returns",
            "Compare which scenario favors buying vs renting",
        ]),
        ("IMPORTANT NOTES", [
            "This calculator does not include: PMI, moving costs, transaction costs on sale",
            "Tax savings assume you itemize; many filers take standard deduction",
            "Renter scenario simplifies: invests down payment; doesn't model monthly savings flow",
            "For personalized advice, consult a financial advisor or CPA",
            "© 2026 ClearMetric. For educational use only. Not financial advice.",
        ]),
    ]

    r = 4
    for title, items in sections:
        ws.cell(row=r, column=2, value=title).font = Font(name="Calibri", size=12, bold=True, color=TEAL)
        ws.cell(row=r, column=2).fill = FILL_LIGHT
        ws.cell(row=r, column=2).border = THIN
        r += 1
        for item in items:
            ws.cell(row=r, column=2, value=item).font = Font(name="Calibri", size=11, color="2C3E50")
            ws.cell(row=r, column=2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[r].height = 22
            r += 1
        r += 1


# ============================================================
# MAIN
# ============================================================
def main():
    wb = openpyxl.Workbook()
    ws = wb.active

    print("Building Rent vs Buy Calculator sheet...")
    build_rent_vs_buy(ws)

    print("Building Year-by-Year sheet...")
    build_year_by_year(wb)

    print("Building What-If Scenarios sheet...")
    build_what_if(wb)

    print("Building How To Use sheet...")
    build_instructions(wb)

    wb.active = 0

    out = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                       "output", "ClearMetric-Rent-vs-Buy-Calculator.xlsx")
    os.makedirs(os.path.dirname(out), exist_ok=True)
    wb.save(out)
    print(f"\nSaved: {out}")
    print(f"Size: {os.path.getsize(out) / 1024:.1f} KB")
    print(f"Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
